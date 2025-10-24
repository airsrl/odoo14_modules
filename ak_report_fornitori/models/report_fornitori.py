import os
import base64
import xlwt
from odoo import fields, models, api,_
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import pandas as pd
import math
import pytz
import openpyxl
import re
from odoo import exceptions
import time
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT, pycompat,float_round,format_datetime
import io

import csv
#SI PRENDONO I RECORD COMPLESSIVI DAll'inventario
#AMD
amd_inv_field_list=['reseller_custom','seller_code','description','quantity','qty_on_order','data']
amd_inv_title_list=['Reseller Customer Number','Manufacturer Part Number or OPN','Description',
                'Quantity On Hand','Quantity on order','Date']
#NVIDIA
nvidia_inv_field_list=['reporter_id','data','barcode','marchio_id','description','quantity','qty_in_po_order']
nvidia_inv_title_list=['Reporter ID','Inventory Date','EAN','Manufacturer',
                'Product Description','Quantity on Hand','PO Back Log']
#MSI
msi_inv_field_list=['stock_move_id','default_code','barcode','description','quantity','cost_ultimo']
msi_inv_title_list=['Mag','Codice Articolo','Secondo Codice',
                'Descrizione','Giacenza','Costo ultimo']
#ASUS
asus_inv_field_list = ['asus_default_code', 'qty_start_week', 'qty_new_coming', 'qty_sold', 'qty_end_week',
                       'qty_new_order', '', '', '', '', '', 'qty_returned', 'qty_damaged', 'part_number', '', '', '',
                       '', '', 'default_code', '', '', 'qty_preorder', 'qty_to_process']
asus_inv_title_list = ['MODEL', 'INITIAL_INVENTORY', 'NEW_COMING', 'SELLOUT', 'FINAL_INVENTORY', 'NEW_ORDER_QTY',
                       'REMARKS', 'DESCRIPTION', 'TRANSIT', 'TRANSIT_QTY', 'CONSIGN_QTY', 'OTHER_NEW_COMING',
                       'OTHER_STOCK_OUT', 'PART_NO', 'INVOICE', 'EAN_CODE', 'UPC_CODE', 'MFG_PART_NO', 'MFG_MODEL_NAME',
                       'CUSTOMER_SKU', 'WAREHOUSE_NAME', 'WAREHOUSE_CODE', 'SO_BACKLOG', 'SO_PROCESSING']
#Si PRENDONO I DATI DELLA SETTIMANA DA ACOUNT.MOVE.LINE
#AMD
amd_pos_field_list=['reseller_custom','seller_code','description','qty_sold','invoice_date','invoice_id','product_type','sold_to_name'
                    ,'sold_to_address','sold_to_city','sold_to_state','sold_to_zip','sold_to_country']
amd_pos_title_list=['Reseller Customer Number','Manufacturer Part Number or OPN','Description',
                'Qty sold','Invoice Date','Invoice','POS Type','Sold To Name','Sold To Address','Sold To City'
                ,'Sold To State',
                    'Sold To Zip/Postal Code','Sold To Country']
#NVIDIA
inv_pos_field_list=['reporter_id','barcode','marchio_id','description','qty_sold','invoice_date','invoice_id','price_unit','currency_id','sold_to_name',
                    'sold_to_zip','sold_to_country']
inv_pos_title_list=['Reporter ID','EAN','Manufacturer','Product Description',
                'Quantity Sold/Returned','Sold/Invoice Date','Invoice Number','Price Per Unit','Currency of Sale','Ship To Name',
                    'Ship to Postal Code','Ship To Country Code']
#MSI
msi_pos_field_list=['invoice_date','default_code','barcode','description','sold_to_name'
                    ,'sold_to_vat','city','state_id','name','partner_id','product_type','warehouse_id','qty_sold','price_unit',
                    'categ1','categ2','categ3']
msi_pos_title_list=['Data Movimenti','Codice Articolo','Secondo Codice',
                'Articoli.Descrizione','Ragione sociale','PartitaIVA',	'Località',	'Provincia','NFattura','Magazzino.Descrizione'
                    ,'Causale','Mag','Qta','Prezzo Vendita unit','Cat 1','Cat 2','Cat 3'	]
#SAMSUNG
samsung_title_list=['code','shipto','sellout','stock','ean']
class ReportFornitori(models.Model):
    _name = 'report.fornitori'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _description = 'Description'
    _order= "create_date desc"
    name = fields.Char('Nome report')
    marchio_ids=fields.Many2many('ak_connettore.marchio',string='Marchi')
    partner_id=fields.Many2one('res.partner','Distributore')
    email = fields.Char('Email')
    email_cc =fields.Char('Email CC')
    automatic_send=fields.Boolean('Invia automaticamente')
    categorie_ids = fields.Many2many('product.category',string='Categorie')
    report_line_id = fields.One2many('report.fornitori.line','parent_id','Report Generati')
    report_type_fornitore= fields.Selection([('amd','Amd'),('msi','Msi'),('nvidia','Nvidia'),('samsung','Samsung'),('asus', 'Asus')],  'Report di riferimento')
    date_start=fields.Date()
    date_stop=fields.Date()
    def get_date_oggi(self):
        oggi = date.today()

        return oggi
    def get_date_from(self):
        return date.today() - relativedelta(days=7)

    def call_action_stock_valuation(self,product_ids,date_stop=False):
        #todo da vedere
        if not date_stop:
            date_stop=self.date_stop
        date_stop=datetime.strptime(date_stop.strftime('%d-%m-%Y 23:59:59'),'%d-%m-%Y %H:%M:%S')
        wiz = self.env["stock.valuation.layer"].with_context({'search_default_group_by_product_id': 1}).search([('create_date', '<=', date_stop),('product_id','in',product_ids.ids),('quantity','!=',0)])


        return wiz
    def last_cost_stock_valuation(self,product_ids,date_stop=False):
        #todo da vedere
        if not date_stop:
            date_stop=self.date_stop
        date_stop=datetime.strptime(date_stop.strftime('%d-%m-%Y 23:59:59'),'%d-%m-%Y %H:%M:%S')
        wiz = self.env["stock.move.line"].with_context({'search_default_group_by_product_id': 1}).search([('state','=','done'),('date', '<=', date_stop),('product_id','in',product_ids.ids),('in_out_move','>',0),('movimenti_prodotto','>',0)],order='date asc')


        return wiz

    #Funzione ritorna la product list
    def return_product_list(self):
        product_product = self.sudo().env['product.product']
        search_domain = []
        if self.marchio_ids:
            search_domain.append(('marchio_id', 'in', self.marchio_ids.ids))
        if self.categorie_ids:
            search_domain.extend(
                ['|', ('categ_id', 'in', self.categorie_ids.ids), ('categ_id', 'child_of', self.categorie_ids.ids)])

        product_template = self.sudo().env['product.template'].search(search_domain)
        product_list = product_product.search([('product_tmpl_id', 'in', product_template.ids)])
        return product_list

    def create_sell_xls(self):
        product_list = self.return_product_list()

        if product_list:
            if self.report_type_fornitore == 'amd':

                self.generate_xls_amd_attachment_pos(product_list=product_list)
            elif self.report_type_fornitore =='nvidia':

                self.generate_xls_nv_attachment_pos(product_list=product_list)
            elif self.report_type_fornitore =='msi':

                self.generate_xls_msi_attachment_pos(product_list=product_list)
    def create_stock_xls(self):
        product_list = self.return_product_list()

        giacenze_line = self.call_action_stock_valuation(product_list)
        if product_list:
            if self.report_type_fornitore == 'amd':
                self.generate_xls_amd_attachment_stock(lines=giacenze_line)

            elif self.report_type_fornitore =='nvidia':
                self.generate_xls_attachment_nv(lines=giacenze_line)

            elif self.report_type_fornitore =='msi':
                self.generate_xls_attachment_msi(lines=giacenze_line)

            elif self.report_type_fornitore == 'asus':
                self.generate_xls_attachment_asus(product_ids=product_list)
    def schedule_invio(self):
        records=self.env['report.fornitori'].search([('automatic_send','=',True)])
        for rec in records:
            draft=rec.create_xls()
            context=draft['context']
            mail_compose_create=self.env[draft['res_model']].with_context(context).create({})
            d= mail_compose_create.action_send_mail()

    def create_xls(self):
        attachment=[]
        product_list = self.return_product_list()
        stock= line_stock =sell = line_sell= False
        giacenze_line = self.call_action_stock_valuation(product_list)
        if product_list:
            if self.report_type_fornitore == 'amd':
                stock,line_stock=self.generate_xls_amd_attachment_stock(lines=giacenze_line)
                sell,line_sell=self.generate_xls_amd_attachment_pos(product_list=product_list)


            elif self.report_type_fornitore =='nvidia':
                stock,line_stock=self.generate_xls_attachment_nv(lines=giacenze_line)
                sell,line_sell=self.generate_xls_nv_attachment_pos(product_list=product_list)
            elif self.report_type_fornitore =='msi':
                #.read()[0]
                stock, line_stock =self.generate_xls_attachment_msi(lines=giacenze_line)
                sell,line_sell=self.generate_xls_msi_attachment_pos(product_list=product_list)
            elif self.report_type_fornitore == 'samsung':
                stock, line_stock = self.generate_csv_samsung(product_list,giacenze_line,lista_titoli=samsung_title_list)
            if stock:
                attachment.append(stock.id)
                line_stock.date = date.today()
            if sell:
                attachment.append(sell.id)
                line_sell.date=date.today()
        template_id = self.env['ir.model.data'].xmlid_to_res_id('ak_report_fornitori.email_report_base',
                                                                raise_if_not_found=False)
        lang = self.env.context.get('lang')
        template = self.sudo().env['mail.template'].browse(template_id)
        if template.lang:
            lang = template._render_lang(self.ids)[self.id]
        ctx = {
            'default_email_to':self.email,
            'default_email_cc':self.email_cc,
            'default_model': 'report.fornitori',
            'default_res_id': self.ids[0],
            'default_use_template': bool(template_id),
            'default_template_id': template_id,
            'default_composition_mode': 'mass_mail',
            'force_email': True,
           'default_attachment_ids': [(6,0,attachment)],
        }
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'views': [(False, 'form')],
            'view_id': False,
            'target': 'new',
            'context': ctx,
        }
   ############################################
   #####           AMD                 ########
   ############################################
    def search_domain_pos(self,prodotti,move_type,date_start=False,date_stop=False,product_type=False):

        invoices=[]
        if not date_start:
            date_start = self.date_start
        if not date_stop:
            date_stop = self.date_stop



        domain = [

            ('invoice_date', '>=', date_start),
            ('invoice_date', '<=', date_stop),
            ('product_id','in',prodotti.ids),
            ('move_type','=',move_type)

        ]
        if product_type:
            domain.append(('product_type','=',product_type))


        invoices = self.env['report.fornitori.data'].search(domain)

        return invoices
    def generate_xls_amd_attachment_pos(self, product_list,lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = amd_pos_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = amd_pos_title_list
        if not lines:


            #lines= self.search_domain_pos(product_list,move_type='invoice')
            lines1 = self.search_domain_pos(product_list, move_type='invoice', product_type='comp')
            lines2 = self.search_domain_pos(product_list, move_type='bom')
            lines = lines1 + lines2

        if not name:
            settimana=self.date_stop.isocalendar()[1]
            name = 'POS(sell)-W%s-AK INFO 90168.xls'%  settimana
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Amd_sell", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)

        riga = 1
        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0

        for line in lines:
            shipping_id = line.partner_id
            for field in lista_campi:
                # Se il campo è presente c lo scrivo
                if field in line:
                    new_field = line[field]

                if field in ['seller_code']:
                    seller_code= line.product_id.default_code
                    sheet.write(riga, lista_campi.index(field), (seller_code if seller_code else ' '))
                elif field in ['reseller_custom']:
                    sheet.write(riga, lista_campi.index(field), '901680')
                elif field in ['description']:
                    description = line.product_id.product_tmpl_id.with_context(lang='it_IT').name
                    sheet.write(riga, lista_campi.index(field), (description if description else ' '))
                elif field in ['qty_sold']:
                    sheet.write(riga, lista_campi.index(field), (line.quantity if line.quantity else 0))
                elif field in ['invoice_date']:
                    sheet.write(riga, lista_campi.index(field), (line.invoice_date.strftime('%m/%d/%Y') if line.invoice_date else ' '))
                    if line.product_type == 'direct':
                        sheet.write(riga, lista_campi.index(field),
                                    (line.mrp_id.date_finished.strftime('%m/%d/%Y') if line.mrp_id.date_finished else ' '))
                elif field in ['invoice_id']:
                    if line.move_id:
                        sheet.write(riga, lista_campi.index(field),  (line.move_id.name if line.move_id else ' '))
                        if line.product_type=='direct':
                            sheet.write(riga, lista_campi.index(field),
                                        (line.mrp_id.display_name if line.mrp_id else ' '))
                    elif line.stock_move:
                        sheet.write(riga, lista_campi.index(field),
                                    (line.stock_move.reference if line.stock_move.reference else line.stock_move.name))

                    elif line.pos_order_id:
                        sheet.write(riga, lista_campi.index(field),  (line.stock_move.name if line.stock_move else ' '))



                elif field in ['sold_to_name']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.name if shipping_id  else '')
                elif field in ['sold_to_address']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.street if shipping_id.street else ' ')
                elif field in ['sold_to_city']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.city if shipping_id.city else ' ')
                elif field in ['sold_to_state']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.state_id.code if shipping_id.state_id else ' ')
                elif field in ['sold_to_zip']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.zip if shipping_id.zip else ' ')
                elif field in ['product_type']:
                    pos_type=''
                    if new_field=='comp':
                        pos_type='Comp'
                    elif new_field=='direct':
                        pos_type='SI'
                    sheet.write(riga, lista_campi.index(field),pos_type )
                elif field in ['sold_to_country']:
                    sheet.write(riga, lista_campi.index(field), shipping_id.country_id.code if shipping_id.country_id else ' ')


            riga += 1

        # Salvo il file
       # wb.save(name)

        wb.save(f)
        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line

    def generate_xls_amd_attachment_stock(self, lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = amd_inv_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = amd_inv_title_list

        if not name:
            settimana=self.date_stop.isocalendar()[1]
            name = 'INV(stock)-W%s-AK INFO 90168.xls'%  settimana
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Amd_stock", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)


        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0

        data_list=[]
        for line in lines:
            data_dict = {}
            for field in lista_campi:
                # Se il campo è presente c lo scrivo
                if field in line:
                    new_field = line[field]
                elif field in line.product_id:
                    new_field = line.product_id[field]

                if field in ['description']:
                    data_dict[field]=line.product_id.product_tmpl_id.with_context(lang='it_IT').name if line.product_id else ' '
                elif field in ['quantity']:
                    data_dict[field]=new_field if new_field else 0
                elif field in ['seller_code']:
                    data_dict[field]=line.product_id.default_code if line.product_id else ' '
                elif field in ['reseller_custom']:
                    data_dict[field]= '901680'
                elif field in ['qty_on_order']:
                    data_dict[field]=0
                elif field in ['data']:
                    data_dict[field]= self.date_stop.strftime('%m/%d/%Y')
            data_list.append(data_dict)


        # Salvo il file
       # wb.save(name)
        df = pd.DataFrame(data_list, columns=lista_campi)
        lista_group=lista_campi.copy()
        lista_group.remove('quantity')
        recordset = df.groupby(lista_group, as_index=False)['quantity'].sum()



        records = recordset.to_dict('records')
        # EXPORT IN EXCEL
        riga = 1
        for rec in records:
            if rec['quantity'] >0:
                for key,value in rec.items():
                    sheet.write(riga, lista_campi.index(key), value)
                riga +=1
        wb.save(f)
        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line
    ############################################
    #####              NVIDIA            #######
    ############################################
    def generate_xls_attachment_nv(self, lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = nvidia_inv_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = nvidia_inv_title_list

        if not name:
            giorno=self.date_stop.strftime('%Y-%m-%d')
            name = 'AK INF-INV-%s.xls'%  giorno
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Q_stock", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)

        riga = 1
        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0
        data_list=[]
        for line in lines:
            data_dict = {}
            for field in lista_campi:
                # Se il campo è presente c lo scrivo
                if field in line:
                    new_field = line[field]

                if field in ['reporter_id']:
                    data_dict[field]=  'AK Informatica'
                elif field in ['quantity']:
                    data_dict[field]=  new_field if new_field else 0
                elif field in ['marchio_id']:
                    data_dict[field]= line.product_id[field].name if line.product_id[field] else ' '
                elif field in ['barcode']:
                    data_dict[field]= line.product_id[field] if line.product_id else ' '
                elif field in ['description']:
                    data_dict[field]=line.product_id.product_tmpl_id.with_context(lang='it_IT').name if line.product_id else ' '
                elif field in ['qty_in_po_order']:
                    #todo in teoria non serve
                    # qty_on_order=line.product_id.get_purchased_product_qty(date_to=self.date_stop)
                    # data_dict[field]=  qty_on_order
                    data_dict[field] =0
                elif field in ['data']:
                    data_dict[field]= self.date_stop.strftime('%d/%m/%Y')

            data_list.append(data_dict)

        # Salvo il file
        # wb.save(name)
        df = pd.DataFrame(data_list, columns=lista_campi)
        lista_group = lista_campi.copy()
        lista_group.remove('quantity')
        recordset = df.groupby(lista_group, as_index=False)['quantity'].sum()

        records = recordset.to_dict('records')
        # EXPORT IN EXCEL
        riga = 1
        for rec in records:
            if rec['quantity'] > 0:
                for key, value in rec.items():
                    sheet.write(riga, lista_campi.index(key), value)
                riga += 1

        wb.save(f)
        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line
    def generate_xls_nv_attachment_pos(self, product_list,lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = inv_pos_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = inv_pos_title_list
        if not lines:
            lines1= self.search_domain_pos(product_list,move_type='invoice', product_type='comp')
            lines2 = self.search_domain_pos(product_list, move_type='bom')
            lines =lines1 + lines2
        if not name:
            giorno=self.date_stop.strftime('%Y-%m-%d')
            name = 'SO-AK INFO-%s.xls'%  giorno
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Q_Stat_Venduto_Nvidia_Trova1", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)

        riga = 1
        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0

        for line in lines:
            shipping_id = line['partner_id']
            for field in lista_campi:
                # Se il campo è presente c lo scrivo
                if field in line:
                    new_field = line[field]

                if field in ['reporter_id']:
                    sheet.write(riga, lista_campi.index(field), 'Ak Informatica')
                elif field in ['barcode']:
                    default_code=line.product_id.barcode if line.product_id else ' '
                    sheet.write(riga, lista_campi.index(field), (default_code if default_code else ' ') )
                elif field in ['description']:
                    description = line.product_id.product_tmpl_id.with_context(lang='it_IT').name if line.product_id else ' '
                    sheet.write(riga, lista_campi.index(field), description )
                elif field in ['qty_sold']:
                    sheet.write(riga, lista_campi.index(field), (line.quantity if line.quantity else 0))
                elif field in ['invoice_date']:
                    if line['product_type'] == 'direct':
                        sheet.write(riga, lista_campi.index(field),
                                    (line.mrp_id.date_finished.strftime('%d/%m/%Y') if  line.mrp_id else line.stock_move.date.strftime('%d/%m/%Y')))
                    else:
                        if line.move_id:
                            sheet.write(riga, lista_campi.index(field), (line.move_id.date.strftime('%d/%m/%Y') if line['move_id'].date else ' '))

                        elif line.stock_move:
                            sheet.write(riga, lista_campi.index(field),
                                        (line.stock_move.date.strftime('%d/%m/%Y') if line.stock_move else ' '))
                        elif line.invoice_date:
                            sheet.write(riga, lista_campi.index(field),
                                        (line.invoice_date.strftime('%d/%m/%Y') if line.invoice_date else ' '))
                elif field in ['invoice_id']:

                    if line['product_type']=='direct':
                            sheet.write(riga, lista_campi.index(field),
                                        (line.mrp_id.display_name if line.mrp_id else ' '))
                    elif line.stock_move_line:
                        sheet.write(riga, lista_campi.index(field),
                                    (line.stock_move_line.reference if line.stock_move_line else ' '))
                    else:
                        sheet.write(riga, lista_campi.index(field), (line['move_id'].name if line['move_id'] else ' '))
                elif field in ['sold_to_name']:
                    sheet.write(riga, lista_campi.index(field), ('PC Production' if line.product_type == 'direct'  else 'Anonimous'))
                elif field in ['marchio_id']:
                    marchio_id =line['product_id'].marchio_id
                    sheet.write(riga, lista_campi.index(field), (marchio_id.name if marchio_id else ' '))
                elif field in ['price_unit']:
                    if line.product_type == 'direct':
                        date_stop=line.mrp_id.date_finished if line.mrp_id else line.stock_move
                        costo_ultimo = self.last_cost_stock_valuation(line.product_id,date_stop)
                        sheet.write(riga, lista_campi.index(field), costo_ultimo[-1].movimenti_prodotto if costo_ultimo else 0)

                    else:
                        if line.move_line:
                            sheet.write(riga, lista_campi.index(field), line['move_line'][0][field] if line['move_line'] else 0)
                        elif line.pos_order_id:
                            price_unit,price_total=self.get_price_posorder(line.pos_order_id,line.product_id)
                            sheet.write(riga, lista_campi.index(field),
                                        price_unit if price_unit else 0)
                        else:
                            date_stop = line.invoice_date if  line.invoice_date else False
                            costo_ultimo = self.last_cost_stock_valuation(line.product_id,date_stop)
                            sheet.write(riga, lista_campi.index(field), costo_ultimo[-1].movimenti_prodotto if costo_ultimo else 0)
                elif field in ['currency_id']:
                    if line.move_id:
                        currency_id = line['move_id'].currency_id
                        sheet.write(riga, lista_campi.index(field), (currency_id.name if currency_id else ' '))
                    elif line.pos_order_id:
                        currency_id = line.pos_order_id.currency_id
                        sheet.write(riga, lista_campi.index(field), (currency_id.name if currency_id else ' '))
                    else:
                        sheet.write(riga, lista_campi.index(field), 'EUR')
                elif field in ['sold_to_zip']:
                    sheet.write(riga, lista_campi.index(field), ('XXX' if line.product_type == 'direct'  else shipping_id.zip if shipping_id.zip else ' ' ))
                elif field in ['sold_to_country']:
                    sheet.write(riga, lista_campi.index(field), (shipping_id.country_id.code if shipping_id.country_id else ' '))

            riga += 1

        # Salvo il file
       # wb.save(name)

        wb.save(f)
        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line

    ############################################
    #####              Asus               ######
    ############################################
    def generate_xls_attachment_asus(self, product_ids=None, lista_campi=None, name=None, lista_titoli=None,
                                     search_domain=None, scheduled=False):
        """ we need to call the stock evaluation inside this function because we need it both at the start
        and at the end of the current week"""
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = asus_inv_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = asus_inv_title_list

        date_start = self.date_stop - relativedelta(days=7)
        week = self.date_stop.isocalendar()[1]
        company_identification = 'IT-ID-AKINFORMATICA'
        if not name:
            name = 'INVENTORY_%s_Y%s_W%s.xls' % (company_identification, self.date_stop.year, str(week))
        else:
            name = name + '.xls'

        file = io.BytesIO()
        file.name = name

        workbook = xlwt.Workbook(encoding='utf8')
        try:
            # Add a new sheet to the Excel
            sheet = workbook.add_sheet("Inventory", cell_overwrite_ok=True)
        except:
            # Otherwise take the first
            sheet = workbook.get_sheet(0)

        styles = {
            'header_blue': xlwt.easyxf('font: name Times New Roman, color white; pattern: pattern solid, fore_color dark_blue;'),
            'header_grey': xlwt.easyxf('font: name Times New Roman, color black; pattern: pattern solid, fore_color gray25;'),
            'header_grey_centered': xlwt.easyxf('font: name Times New Roman, color black, bold on; pattern: pattern solid, fore_color gray25; align: vert centre, horiz center'),
        }

        # write the document header
        sheet.write_merge(1, 1, 0, 1, 'ASUS Weekly Inventory Report', styles['header_blue'])
        # company name
        sheet.write(2, 0, 'CompanyName', styles['header_grey'])
        sheet.write(2, 1, company_identification)
        # application date
        sheet.write(3, 0, 'ApplicationDate', styles['header_grey'])
        sheet.write(3, 1, str(self.date_stop.year) + 'W' + str(week))
        # method
        sheet.write(4, 0, 'Method', styles['header_grey'])
        sheet.write(4, 1, 'ALL')

        # write the table headers
        # note
        sheet.write_merge(8, 8, 0, 23, 'This file has to be sent every Monday: if not, any price protection right won\'t be valid anymore.', styles['header_blue'])
        # titles
        for title in lista_titoli:
            sheet.write(9, lista_titoli.index(title), title, styles['header_grey_centered'])

        # stock evaluation at the start of the week
        stock_start = self.env["stock.valuation.layer"].read_group([('create_date', '<=', date_start),
                                                                    ('product_id', 'in', product_ids.ids),
                                                                    ('quantity', '!=', 0)],
                                                                   ['product_id', 'quantity:sum'], ['product_id'])
        stock_end = self.env["stock.valuation.layer"].read_group([('create_date', '<=', self.date_stop),
                                                                  ('product_id', 'in', product_ids.ids),
                                                                  ('quantity', '!=', 0)],
                                                                 ['product_id', 'quantity:sum'], ['product_id'])

        product_ids_start = [s['product_id'][0] for s in stock_start]
        product_ids_end = [st['product_id'][0] for st in stock_end]
        # products are all the ones that are either in the stock at the start or at the end of the week
        stock_product_ids = list(set(product_ids_start) | set(product_ids_end))
        stock_product_ids = self.env['product.product'].browse(stock_product_ids)

        row = 10
        for product in stock_product_ids:
            data_is_not_zero = False
            for field in lista_campi:
                if field in ['asus_default_code']:
                    sheet.write(row, lista_campi.index(field), product.default_code_import if product.default_code_import else '')
                elif field in ['qty_start_week']:
                    product_stock_start = [s for s in stock_start if s['product_id'][0] == product.id]
                    product_qty_in_stock_start = product_stock_start[0]['quantity'] if product_stock_start else 0
                    if not data_is_not_zero:
                        data_is_not_zero = bool(product_qty_in_stock_start)
                    sheet.write(row, lista_campi.index(field), product_qty_in_stock_start)
                elif field in ['qty_new_coming']:
                    newcoming_picking_type_ids = self.env['stock.picking.type'].search(
                        [('is_new_coming_picking_type', '=', True)])
                    newcoming_move_ids = self.env['stock.move'].search(
                        [('picking_type_id', 'in', newcoming_picking_type_ids.ids),
                         ('state', '=', 'done'),
                         ('date', '>=', date_start),
                         ('date', '<=', self.date_stop), ('product_id', '=', product.id)])
                    qty_new_coming = sum([line.product_uom_qty for line in newcoming_move_ids]) if newcoming_move_ids else 0
                    if not data_is_not_zero:
                        data_is_not_zero = bool(qty_new_coming)
                    sheet.write(row, lista_campi.index(field), qty_new_coming)
                elif field in ['qty_sold']:
                    sellout_picking_type_ids = self.env['stock.picking.type'].search(
                        [('is_sellout_picking_type', '=', True)])
                    sellout_move_ids = self.env['stock.move'].search(
                        [('picking_type_id', 'in', sellout_picking_type_ids.ids),
                         ('state', '=', 'done'),
                         ('date', '>=', date_start),
                         ('date', '<=', self.date_stop),
                         ('product_id', '=', product.id)])
                    qty_sold = sum([line.product_uom_qty for line in sellout_move_ids]) if sellout_move_ids else 0
                    if not data_is_not_zero:
                        data_is_not_zero = bool(qty_sold)
                    sheet.write(row, lista_campi.index(field), qty_sold)
                elif field in ['qty_end_week']:
                    product_stock_end = [s for s in stock_end if s['product_id'][0] == product.id]
                    product_qty_in_stock_end = product_stock_end[0]['quantity'] if product_stock_end else 0
                    if not data_is_not_zero:
                        data_is_not_zero = bool(product_qty_in_stock_end)
                    sheet.write(row, lista_campi.index(field), product_qty_in_stock_end)
                elif field in ['qty_new_order']:
                    purchase_order_ids = self.env['purchase.order'].search([('state', '=', 'purchase')])
                    purchase_order_lines = self.env['purchase.order.line'].search([
                        ('order_id', 'in', purchase_order_ids.ids), ('product_id', 'in', product.ids)])
                    qty_received = sum([line.qty_received for line in purchase_order_lines]) if purchase_order_lines else 0
                    qty_ordered = sum([line.product_qty for line in purchase_order_lines]) if purchase_order_lines else 0
                    if not data_is_not_zero:
                        data_is_not_zero = bool(qty_ordered - qty_received)
                    sheet.write(row, lista_campi.index(field), qty_ordered - qty_received)
                elif field in ['qty_returned']:
                    sheet.write(row, lista_campi.index(field), 0)
                elif field in ['qty_damaged']:
                    scrap_picking_type_ids = self.env['stock.picking.type'].search(
                        [('is_scrap_picking_type', '=', True)])
                    scrap_move_ids = self.env['stock.move'].search(
                        [('picking_type_id', 'in', scrap_picking_type_ids.ids),
                         ('state', '=', 'done'),
                         ('date', '>=', date_start),
                         ('date', '<=', self.date_stop), ('product_id', '=', product.id)])
                    qty_scrapped = sum([line.product_uom_qty for line in scrap_move_ids]) if scrap_move_ids else 0
                    sheet.write(row, lista_campi.index(field), qty_scrapped)
                elif field in ['part_number']:
                    sheet.write(row, lista_campi.index(field), product.default_code)
                elif field in ['default_code']:
                    sheet.write(row, lista_campi.index(field), product.barcode if product.barcode else product.default_code)
                elif field in ['qty_preorder']:
                    sale_draft_order_line_ids = self.env['sale.order.line'].search([('order_id.state', 'in', ['sale']),
                                                                                    ('product_id', '=', product.id)])
                    picking_ids = sale_draft_order_line_ids.order_id.picking_ids.filtered(lambda p: p.picking_type_id.code in ['outgoing', 'mrp_operation'] and p.state != 'cancel')
                    qty_preorder = sum([line.product_uom_qty - line.reserved_availability - line.quantity_done for line in picking_ids.move_ids_without_package if line.product_id.id == product.id]) if picking_ids else 0
                    sheet.write(row, lista_campi.index(field), qty_preorder)
                elif field in ['qty_to_process']:
                    sale_order_line_not_sent_ids = self.env['sale.order.line'].search(
                        [('order_id.state', 'in', ['sale']), ('product_id', '=', product.id)])
                    picking_ids = sale_order_line_not_sent_ids.order_id.picking_ids.filtered(lambda p: p.picking_type_id.code in ['outgoing', 'mrp_operation'] and p.state != 'cancel')
                    qty_to_process = sum([line.reserved_availability for line in picking_ids.move_ids_without_package if line.product_id.id == product.id]) if picking_ids else 0
                    sheet.write(row, lista_campi.index(field), qty_to_process)

            if data_is_not_zero:
                # update the row only if the previous was not empty, otherwise overwrite it
                row += 1
            else:
                # if the last row is empty, it won't be overwritten so fill it with blanks
                for field in lista_campi:
                    sheet.write(row, lista_campi.index(field), '')

        # save the file
        workbook.save(file)
        x = file.seek(0)
        filedata = file.read()
        output = base64.encodebytes(filedata)
        # create_line
        line = self.sudo().env['report.fornitori.line'].create({'name': name, 'file': output, 'parent_id': self.id})
        attachment = self.env['ir.attachment'].create({'name': name,
                                                       'datas': output,
                                                       'res_model': self._name,
                                                       'type': 'binary'})
        return attachment, line

    ############################################
    #####              MSI               #######
    ############################################
    def generate_xls_attachment_msi(self, lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = msi_inv_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = msi_inv_title_list

        if not name:
            giorno=self.date_stop.strftime('%Y-%m-%d')
            name = 'Stock-MSI-%s.xls'%  giorno
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Q_stock", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)
        #

        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0
        data_list=[]

        for line in lines:
            warehouse_id = False
            data_dict = {}
            for field in lista_campi:
                # Se il campo è presente c lo scrivo

                if field in line:
                    new_field = line[field]

                if field in ['default_code']:
                    data_dict[field]= line.product_id.default_code if line.product_id else ''
                elif field in ['quantity']:
                    if line.product_id.default_code == '911-7D17-05S':
                        c=0
                    data_dict[field] = new_field if new_field else 0
                elif field in ['barcode']:
                    data_dict[field] = line.product_id.barcode if line.product_id.barcode else ''
                elif field in ['cost_ultimo']:
                    date_stop = self.date_stop
                    costo_ultimo= self.last_cost_stock_valuation(line.product_id,date_stop)
                    data_dict[field] = costo_ultimo[-1].movimenti_prodotto if costo_ultimo else 0
                elif field in ['description']:
                    description =  line.product_id.with_context(lang='it_IT').name
                    data_dict[field] = description if description else ''
                elif field in ['stock_move_id']:
                    warehouse_id = new_field.location_id.get_warehouse()
                    if  new_field.location_dest_id.usage=='internal' and line.quantity > 0:
                        warehouse_id = new_field.location_dest_id.get_warehouse()
                    data_dict[field] = warehouse_id.code_msi if warehouse_id else ''
            data_list.append(data_dict)
        lista_group = lista_campi.copy()
        lista_group.remove('quantity')
        df = pd.DataFrame(data_list,columns=lista_campi)
        recordset = df.groupby(lista_group, as_index=False)['quantity'].sum()



        records = recordset.to_dict('records')
        # EXPORT IN EXCEL
        riga = 1
        for rec in records:
            if rec['quantity'] >0:
                codice_prodotto = rec['default_code']
                elem_neg = [i for i in records if i['default_code']==codice_prodotto and i['quantity'] < 0]
                if elem_neg:
                    for el in elem_neg:
                        rec['quantity'] += el['quantity']
                if rec['quantity'] <= 0 :
                    continue
                for key,value in rec.items():
                    sheet.write(riga, lista_campi.index(key), value)

                riga += 1
        # Salvo il file
        wb.save(f)


        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line
    def generate_xls_msi_attachment_pos(self, product_list,lines=None, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not lista_campi or type(lista_campi) is dict:
            lista_campi = msi_pos_field_list
        if not lista_titoli or type(lista_titoli) is dict:
            lista_titoli = msi_pos_title_list
        if not lines:
            date_start=date(date.today().year, 1, 1) if not self.date_start else self.date_start
            lines1= self.search_domain_pos(product_list,date_start=date_start,move_type='invoice',product_type='comp')
            lines2=self.search_domain_pos(product_list,date_start=date_start,move_type='bom')
            lines = lines1+lines2
            #bom_ids = [rec.bom_id.id for rec in lines]
            #lines3=self.search_domain_pos(product_list,date_start=date_start,move_type='invoice')
            # add_line = lines3.filtered(lambda x:x.bom_id not in bom_ids)
            # if add_line:
            #     lines.append(add_line)
        if not name:
            giorno=self.date_stop.strftime('%Y-%m-%d')
            name = 'Sell-MSI-%s.xls'%  giorno
        else:
            name = name + '.xls'
        f = io.BytesIO()
        f.name=name
        wb = xlwt.Workbook(encoding='utf8')
        try:
            # Aggiungo una pagina al file Excel
            sheet = wb.add_sheet("Stat_Venduto", cell_overwrite_ok=True)

        except:
            # Nel caso esiste già prendo la prima
            sheet = wb.get_sheet(0)

        riga = 1
        bold = xlwt.easyxf('font: bold on')
        style_tot = xlwt.easyxf('font: name Times New Roman, color-index red, bold on')
        money = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
                            num_format_str='€#,####.##')
        style_t = xlwt.easyxf('font: bold on; align: vert centre, horiz center')
        # Scrivo i titoli delle colonne
        for titolo in lista_titoli:
            # sheet.write_merge(0, 2, 0, 5, testata, style_t)
            sheet.write(0, lista_titoli.index(titolo), titolo, style_t)

        # Ciclo i vari ordini selezionati dall'utente
        tot = 0
        tot1 = 0
        tot2 = 0
        invoice=False
        for line in lines:
            if line.quantity != 0:
                shipping_id = line['partner_id']

                categ = line.product_id.categ_id.complete_name if line.product_id.categ_id else ' / / /'
                categ = categ.split('/')
                for field in lista_campi:
                    # Se il campo è presente c lo scrivo
                    if field in line:
                        if field in line['move_line']:
                            invoice =  line['move_line'][field]
                        new_field =  line[field]

                    if field in ['default_code']:
                        default_code=line['product_id'].default_code
                        sheet.write(riga, lista_campi.index(field), (default_code if default_code else ' ') )
                    elif field in ['barcode']:

                        sheet.write(riga, lista_campi.index(field), (line.product_id.barcode if line.product_id.barcode else ' '))
                    elif field in ['description']:
                        description = line.product_id.with_context(lang='it_IT').name
                        sheet.write(riga, lista_campi.index(field), (description if description else ' '))
                    elif field in ['qty_sold']:
                        sheet.write(riga, lista_campi.index(field), (line.quantity if line.quantity else 0))

                    elif field in ['partner_id']:
                        if line['move_line']:
                            if line['move_line'][0].sale_line_ids:
                                warehouse_id =line['move_line'][0].sale_line_ids[0].warehouse_id
                                sheet.write(riga, lista_campi.index(field), (warehouse_id.name if warehouse_id else ' '))
                        if line.product_type == 'direct':
                            warehouse_id = line.stock_move.location_id.get_warehouse()
                            sheet.write(riga, lista_campi.index(field), (warehouse_id.name if warehouse_id else ' '))
                    elif field in ['product_type']:
                        pos_type = ''
                        if new_field == 'direct':
                            pos_type = '203'
                            sheet.write(riga, lista_campi.index('sold_to_name'), 'Scarico produzione')

                            sheet.write(riga, lista_campi.index('invoice_date'),
                                        (line.mrp_id.date_finished.strftime('%d/%m/%Y') if line.mrp_id.date_finished else ' '))
                            sheet.write(riga, lista_campi.index('name'), line.name)

                        elif new_field == 'comp':
                            pos_type = '201'
                            sheet.write(riga, lista_campi.index('invoice_date'),
                                        (line.invoice_date.strftime('%d/%m/%Y') if line.invoice_date else ' '))

                            sheet.write(riga, lista_campi.index('name'), (line['move_id'].name if line['move_id'] else ' '))

                            sheet.write(riga, lista_campi.index('sold_to_name'), (shipping_id.name if shipping_id else ' '))

                            sheet.write(riga, lista_campi.index('sold_to_vat'), (shipping_id.vat if shipping_id.vat else ' '))

                            sheet.write(riga, lista_campi.index('city'), (shipping_id.city if shipping_id.city else ' '))

                            sheet.write(riga, lista_campi.index('state_id'),
                                        (shipping_id.state_id.code if shipping_id.state_id else ' '))
                        sheet.write(riga, lista_campi.index(field), pos_type)
                    elif field in ['warehouse_id']:
                        if line['move_line']:
                            if line['move_line'][0].sale_line_ids:
                                warehouse_id =line['move_line'][0].sale_line_ids[0].warehouse_id
                                sheet.write(riga, lista_campi.index(field), (warehouse_id.code_msi if warehouse_id else ' '))
                        if line.product_type == 'direct':
                            warehouse_id = line.stock_move.location_id.get_warehouse()
                            sheet.write(riga, lista_campi.index(field), (warehouse_id.code_msi if warehouse_id else ' '))

                    elif field in ['price_unit']:
                        if line.product_type == 'direct':
                            date_stop = line.invoice_date if line.invoice_date else False
                            costo_ultimo =  self.last_cost_stock_valuation(line.product_id,date_stop)
                            sheet.write(riga, lista_campi.index(field), costo_ultimo[-1].movimenti_prodotto if costo_ultimo else 0)
                        elif line.product_type == 'comp':
                            sheet.write(riga, lista_campi.index(field), line['move_line'][0][field] if line['move_line'] else 0)
                    elif field in ['cost_ultimo']:
                        date_stop = line.invoice_date if line.invoice_date else False
                        costo_ultimo = self.last_cost_stock_valuation(line.product_id, date_stop)
                        sheet.write(riga, lista_campi.index(field), (costo_ultimo[-1].movimenti_prodotto if costo_ultimo else 0))
                    elif field in ['categ1']:

                        sheet.write(riga, lista_campi.index(field), (categ[0] if categ else ' '))
                    elif field in ['categ2']:
                        sheet.write(riga, lista_campi.index(field), (categ[1] if len(categ)>1 else ' '))
                    elif field in ['categ3']:
                        sheet.write(riga, lista_campi.index(field), (categ[2] if len(categ)>2 else ' '))

                riga += 1

        # Salvo il file
       # wb.save(name)

        wb.save(f)
        x=f.seek(0)
        filedata = f.read()
        out = base64.encodebytes(filedata)
        #create_line
        line=self.sudo().env['report.fornitori.line'].create({'name':name,'file':out,'parent_id':self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}),line

    ############################################################
    ####### SAMSUNG CSV #######################################
    ############################################################
    def generate_csv_samsung(self, product_list,giacenze, lista_campi=None, name=None, lista_titoli=None,
                                search_domain=None, scheduled=False):
        if not name:
            giorno = self.date_stop
            name = f'Sell-stock-samsung-{giorno}.csv'
        else:
            name = name + '.csv'

        content =  io.BytesIO()

        content.name = name
        lines1 = self.search_domain_pos(product_list, move_type='invoice', product_type='comp')
        lines2 = self.search_domain_pos(product_list, move_type='bom')
        lines = lines1 + lines2
        data_list = []
        #CICLO  I VENDUTI
        for line in lines:
            data_dict =dict.fromkeys(lista_titoli,0)
            data_dict['code']=line.product_id.default_code
            data_dict['ean'] = f'{line.product_id.barcode}'
            if line.move_line:
                if line.move_line[0].sale_line_ids:
                    warehouse_id = line['move_line'][0].sale_line_ids[0].warehouse_id
                    data_dict['shipto']=warehouse_id.code_msi if warehouse_id else ' '
            if line.product_type == 'direct':
                warehouse_id = line.stock_move.location_id.get_warehouse()
                data_dict['shipto']=warehouse_id.code_msi if warehouse_id else ' '
            data_dict['sellout'] = line.quantity
            data_list.append(data_dict)

        #CICLO LE GIACENZE
        for giac in giacenze:

            data_dict =dict.fromkeys(lista_titoli,0)
            data_dict['ean'] = f'{giac.product_id.barcode}'
            data_dict['code'] = giac.product_id.default_code
            data_dict['stock'] = giac.quantity
            warehouse_id = giac.stock_move_id.location_id.get_warehouse()
            if giac.stock_move_id.location_dest_id.usage == 'internal' :
                warehouse_id = giac.stock_move_id.location_dest_id.get_warehouse()
            data_dict['shipto'] = warehouse_id.code_msi if warehouse_id else ''
            data_list.append(data_dict)


        df = pd.DataFrame(data_list, columns=lista_titoli)

        df=df.groupby(['code','shipto','ean'], as_index=False)['sellout','stock'].sum()[lista_titoli]
        num = df._get_numeric_data()
        #porto i valori negativi a 0
        num[num < 0] = 0
        #tolgo le righe che hanno valore sellout e stock a 0
        df = df.drop(df[(df['sellout'] == 0) & (df['stock'] == 0)].index)

        df.to_csv(content,index=False)
        x = content.seek(0)
        filedata = content.read()
        out = base64.encodebytes(filedata)
        # create_line
        line = self.sudo().env['report.fornitori.line'].create({'name': name, 'file': out, 'parent_id': self.id})
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': out,

            'res_model': self._name,
            # 'res_id': ,
            'type': 'binary'}), line
    ################################################
    #                REPORT SAMSUNG                #
    ################################################
    # Funzione  a parte per samsung perchè vuole sia dati di vendita e sia di magazzino insieme
    def create_report_samsung(self):
        product_list = self.return_product_list()
        giacenze_line = self.call_action_stock_valuation(product_list)
        report = self.generate_csv_samsung(product_list,giacenze_line,lista_titoli=samsung_title_list)
    def get_price_posorder(self, pos_order, product):
        lines = pos_order.mapped('lines')
        line_prod = lines.filtered(lambda line: line.product_id == product)

        return line_prod[0].price_unit, line_prod[0].price_subtotal

    def _csv_join(self, *items):
        # Create a csv string for use with data validation formulas and lists.

        # Convert non string types to string.
        items = [str(item) if not isinstance(item, str_types) else item
                 for item in items]

        return ';'.join(items)
class ReportFornitoriLine(models.Model):
    _name='report.fornitori.line'
    _order = 'create_date desc'
    name= fields.Char('Nome file')
    file=fields.Binary('File', copy=False,attachment=True)
    date=fields.Date('Data invio')
    parent_id=fields.Many2one('report.fornitori')

class Product(models.Model):
    _inherit='product.product'
    #todo-fix
    def get_purchased_product_qty(self,date_to=False):
        product=self
        domain = [
            ('order_id.state', 'in', ['purchase']),
            ('product_id', 'in', [product.id]),


        ]
        if date_to:
            date_to = datetime.strptime(date_to.strftime('%d-%m-%Y 23:59:59'), '%d-%m-%Y %H:%M:%S')
            domain.append(('order_id.date_approve', '<=', date_to))


        order_lines = self.env['purchase.order.line'].read_group(domain, ['product_id', 'product_uom_qty'], ['product_id'])
        purchased_data = dict([(data['product_id'][0], data['product_uom_qty']) for data in order_lines])
        purchased_product_qty = float_round(purchased_data.get(product.id, 0), precision_rounding=product.uom_id.rounding)
        return purchased_product_qty

